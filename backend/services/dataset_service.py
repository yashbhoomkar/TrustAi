import logging
import os
import shutil
import uuid

from pathlib import Path

from fastapi import UploadFile

from openpyxl import load_workbook

from database.postgres.connection import (
    SessionLocal
)

from database.postgres.dataset_repository import (
    create_dataset,
    get_datasets,
    get_dataset,
    update_dataset,
    delete_dataset
)

logger = logging.getLogger(__name__)


###########################################################
# Storage Folder
###########################################################

DATASET_STORAGE = Path(
    "storage/datasets"
)

DATASET_STORAGE.mkdir(
    parents=True,
    exist_ok=True
)

SUPPORTED_EXTENSIONS = {
    ".xlsx"
}


###########################################################
# Upload Dataset
###########################################################

def upload_dataset(

    user_id: int,

    display_name: str,

    file: UploadFile

):

    logger.info(
        f"Uploading dataset for User {user_id}"
    )

    extension = Path(
        file.filename
    ).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:

        return {

            "status": "error",

            "message": "Only .xlsx files are supported."

        }

    unique_filename = (
        f"{uuid.uuid4()}{extension}"
    )

    user_folder = (
        DATASET_STORAGE /
        str(user_id)
    )

    user_folder.mkdir(
        parents=True,
        exist_ok=True
    )

    storage_path = (
        user_folder /
        unique_filename
    )

    with open(
        storage_path,
        "wb"
    ) as buffer:

        shutil.copyfileobj(
            file.file,
            buffer
        )

    ###########################################################
    # Read Excel
    ###########################################################

    workbook = load_workbook(
        storage_path,
        data_only=True
    )

    worksheet = workbook.active

    headers = []

    for cell in worksheet[1]:

        headers.append(

            str(cell.value)

            if cell.value is not None

            else ""

        )

    column_count = len(headers)

    row_count = max(

        worksheet.max_row - 1,

        0

    )

    ###########################################################

    file_size = os.path.getsize(
        storage_path
    )

    db = SessionLocal()

    try:

        dataset = create_dataset(

            db=db,

            user_id=user_id,

            display_name=display_name,

            original_filename=file.filename,

            stored_filename=unique_filename,

            storage_path=str(storage_path),

            file_size=file_size,

            rows=row_count,

            columns=column_count,

            column_names=headers,

            column_mapping={}

        )

        return {

            "status": "success",

            "dataset_id": dataset.id,

            "message": "Dataset uploaded successfully."

        }

    finally:

        db.close()


###########################################################
# List Datasets
###########################################################

def list_datasets(
    user_id: int
):

    db = SessionLocal()

    try:

        datasets = get_datasets(
            db,
            user_id
        )

        response = []

        for dataset in datasets:

            response.append({

                "id": dataset.id,

                "display_name": dataset.display_name,

                "original_filename": dataset.original_filename,

                "status": dataset.status,

                "rows": dataset.rows,

                "columns": dataset.columns,

                "column_names": dataset.column_names,

                "column_mapping": dataset.column_mapping,

                "created_at": str(
                    dataset.created_at
                )

            })

        return response

    finally:

        db.close()


###########################################################
# Get Dataset
###########################################################

def get_dataset_details(

    user_id: int,

    dataset_id: int

):

    db = SessionLocal()

    try:

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        return dataset

    finally:

        db.close()


###########################################################
# Preview Dataset
###########################################################

def preview_dataset(

    user_id: int,

    dataset_id: int

):

    db = SessionLocal()

    try:

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        workbook = load_workbook(

            dataset.storage_path,

            data_only=True

        )

        worksheet = workbook.active

        rows = list(

            worksheet.iter_rows(
                values_only=True
            )

        )

        if len(rows) == 0:

            return {

                "columns": [],

                "mapping": dataset.column_mapping,

                "rows": []

            }

        columns = [

            str(value)

            if value is not None

            else ""

            for value in rows[0]

        ]

        preview_rows = [

            list(row)

            for row in rows[1:101]

        ]

        return {

            "columns": columns,

            "mapping": dataset.column_mapping,

            "rows": preview_rows

        }

    finally:

        db.close()


###########################################################
# Update Column Mapping
###########################################################

def update_dataset_mapping(

    user_id: int,

    dataset_id: int,

    mapping: dict

):

    db = SessionLocal()

    try:

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        update_dataset(

            db,

            dataset,

            column_mapping=mapping

        )

        return {

            "status": "success",

            "message": "Column mapping saved."

        }

    finally:

        db.close()


###########################################################
# Delete Dataset
###########################################################

def remove_dataset(

    user_id: int,

    dataset_id: int

):

    db = SessionLocal()

    try:

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        if os.path.exists(
            dataset.storage_path
        ):

            os.remove(
                dataset.storage_path
            )

        delete_dataset(

            db,

            dataset

        )

        return {

            "status": "success",

            "message": "Dataset deleted successfully"

        }

    finally:

        db.close()

###########################################################
# Update Column Mapping
###########################################################

def update_dataset_mapping(

    user_id: int,

    dataset_id: int,

    mapping: dict

):

    db = SessionLocal()

    try:

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        update_dataset(

            db,

            dataset,

            column_mapping=mapping

        )

        return {

            "status": "success",

            "message": "Column mapping saved"

        }

    finally:

        db.close()