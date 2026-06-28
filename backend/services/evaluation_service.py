import json
import logging
import shutil

from pathlib import Path

from openpyxl import load_workbook
from pathlib import Path
from fastapi.responses import FileResponse

from database.postgres.connection import (
    SessionLocal
)

from database.postgres.dataset_repository import (
    get_dataset
)

from database.postgres.metric_repository import (
    get_metric
)

from database.postgres.evaluation_repository import (

    create_evaluation,

    get_evaluations,

    get_evaluation,

    update_evaluation,

    delete_evaluation

)

from services.gemini_service import (
    evaluate_row
)

logger = logging.getLogger(__name__)


###########################################################
# Storage
###########################################################

EVALUATION_STORAGE = Path(
    "storage/evaluations"
)

EVALUATION_STORAGE.mkdir(
    parents=True,
    exist_ok=True
)


###########################################################
# Create Evaluation
###########################################################

def create_new_evaluation(

    user_id: int,

    evaluation_name: str,

    dataset_id: int,

    metric_ids: list[int]

):

    print("=== STEP 1 === ")
    logger.info(
        f"Creating evaluation '{evaluation_name}'"
    )

    db = SessionLocal()

    try:

        ###################################################
        # Load Dataset
        ###################################################

        dataset = get_dataset(

            db,

            user_id,

            dataset_id

        )

        print("=== STEP 2 ===")
        print(dataset.column_mapping)

        if dataset is None:

            return {

                "status": "error",

                "message": "Dataset not found"

            }

        logger.info(

            f"Dataset loaded ({dataset.id})"

        )

        ###################################################
        # Validate Mapping
        ###################################################

        if not dataset.column_mapping:

            return {

                "status": "error",

                "message": "Dataset mapping not configured"

            }

        mapping = dataset.column_mapping

        required_fields = [
            "User Prompt",
            "Expected Response",
            "LLM Response"
        ]

        for field in required_fields:

            if field not in mapping:

                return {
                    "status": "error",
                    "message": f"{field} mapping is missing."
                }

        ###################################################
        # Read Workbook
        ###################################################

        workbook = load_workbook(

            dataset.storage_path,

            data_only=True

        )

        worksheet = workbook.active

        total_rows = max(

            worksheet.max_row - 1,

            0

        )

        workbook.close()
        print("==== STEP 3 ==== ")

        logger.info(

            f"{total_rows} rows detected"

        )

        ###################################################
        # Load Metrics
        ###################################################

        metric_snapshots = []

        for metric_id in metric_ids:

            metric = get_metric(

                db,

                user_id,

                metric_id

            )

            if metric is None:

                return {

                    "status":"error",

                    "message":f"Metric {metric_id} not found"

                }

            metric_snapshots.append({

                "id": metric.id,

                "title": metric.title,

                "description": metric.description,

                "system_prompt": metric.system_prompt,

                "general_instructions": metric.general_instructions,

                "output_type": metric.output_type,

                "min_value": metric.min_value,

                "max_value": metric.max_value,

                "discrete_values": metric.discrete_values,

                "is_default": metric.is_default

            })

        logger.info(

            f"{len(metric_snapshots)} metrics loaded"

        )

        ###################################################
        # Create Evaluation
        ###################################################

        evaluation = create_evaluation(

            db=db,

            user_id=user_id,

            dataset_id=dataset_id,

            evaluation_name=evaluation_name,

            metrics=metric_snapshots,

            total_rows=total_rows,

            results_path=""

        )

        logger.info(

            f"Evaluation {evaluation.id} created"

        )

        ###################################################
        # Create Results File
        ###################################################

        results_path = (

            EVALUATION_STORAGE /

            f"evaluation_{evaluation.id}.xlsx"

        )

        update_evaluation(

            db,

            evaluation,

            results_path=str(results_path)

        )

        ###################################################
        # Run Evaluation
        ###################################################

        run_evaluation(

            db,

            evaluation,

            dataset

        )

        return {

            "status":"success",

            "evaluation_id":evaluation.id

        }

    finally:

        db.close()


###########################################################
# Run Evaluation
###########################################################

def run_evaluation(

    db,

    evaluation,

    dataset

):

    logger.info(

        f"Running Evaluation {evaluation.id}"

    )

    ###################################################
    # Mark Running
    ###################################################

    update_evaluation(

        db,

        evaluation,

        status="running",

        completed_rows=0

    )

    ###################################################
    # Create Evaluation Workbook
    ###################################################

    shutil.copyfile(

        dataset.storage_path,

        evaluation.results_path

    )

    workbook = load_workbook(

        evaluation.results_path

    )

    worksheet = workbook.active

    rows = list(

        worksheet.iter_rows(

            values_only=True

        )

    )

    ###################################################
    # Read Headers
    ###################################################

    headers = [

        str(value).strip()

        if value is not None

        else ""

        for value in rows[0]

    ]

    ###################################################
    # Column Mapping
    ###################################################

    mapping = dataset.column_mapping

    logger.info(
        f"Resolved Mapping : {mapping}"
    )

    ###################################################
    # Create Metric Columns
    ###################################################

    metric_columns = {}

    next_column = worksheet.max_column + 1

    for metric in evaluation.metrics:

        worksheet.cell(

            row=1,

            column=next_column

        ).value = f"{metric['title']} Score"

        worksheet.cell(

            row=1,

            column=next_column + 1

        ).value = f"{metric['title']} Reason"

        metric_columns[

            metric["title"]

        ] = (

            next_column,

            next_column + 1

        )

        next_column += 2

    ###################################################
    # Evaluate Rows
    ###################################################

    for index, row in enumerate(

        rows[1:],

        start=2

    ):

        row_dict = dict(

            zip(

                headers,

                row

            )

        )

        ###################################################
        # Dataset Values
        ###################################################

        user_prompt = str(

            row_dict.get(

                mapping["User Prompt"],

                ""

            )

        )

        print("===== MAPPING =====")
        print(mapping)
        print(mapping.keys())

        expected_response = str(

            row_dict.get(

                mapping["Expected Response"],

                ""

            )

        )

        llm_response = str(

            row_dict.get(

                mapping["LLM Response"],

                ""

            )

        )

        logger.info(

            f"Evaluating Excel Row {index}"

        )

        ###################################################
        # Evaluate using Gemini/Ollama
        ###################################################

        metric_results = evaluate_row(

            evaluation.metrics,

            user_prompt,

            expected_response,

            llm_response

        )

        ###################################################
        # Write Results
        ###################################################

        for metric in evaluation.metrics:

            title = metric["title"]

            score_col, reason_col = metric_columns[title]

            value = metric_results.get(

                title,

                {}

            )

            worksheet.cell(

                row=index,

                column=score_col

            ).value = value.get(

                "score"

            )

            worksheet.cell(

                row=index,

                column=reason_col

            ).value = value.get(

                "reason"

            )

        ###################################################
        # Progress
        ###################################################

        update_evaluation(

            db,

            evaluation,

            completed_rows=index - 1

        )

    ###################################################
    # Save Workbook
    ###################################################

    workbook.save(

        evaluation.results_path

    )

    workbook.close()

    ###################################################
    # Finish
    ###################################################

    update_evaluation(

        db,

        evaluation,

        status="completed",

        completed_rows=evaluation.total_rows

    )

    logger.info(

        f"Evaluation {evaluation.id} completed"

    )

###########################################################
# List Evaluations
###########################################################

def list_all_evaluations(user_id: int):

    db = SessionLocal()

    try:

        evaluations = get_evaluations(
            db,
            user_id
        )

        response = []

        for evaluation in evaluations:

            response.append({

                "id": evaluation.id,

                "evaluation_name": evaluation.evaluation_name,

                "status": evaluation.status,

                "total_rows": evaluation.total_rows,

                "completed_rows": evaluation.completed_rows,

                "created_at": str(evaluation.created_at)

            })

        return response

    finally:

        db.close()

###########################################################
# Get Evaluation
###########################################################

def get_evaluation_details(

    user_id: int,

    evaluation_id: int

):

    db = SessionLocal()

    try:

        evaluation = get_evaluation(

            db,

            user_id,

            evaluation_id

        )

        if evaluation is None:

            return {

                "status": "error",

                "message": "Evaluation not found"

            }

        return {

            "id": evaluation.id,

            "evaluation_name": evaluation.evaluation_name,

            "status": evaluation.status,

            "metrics": evaluation.metrics,

            "completed_rows": evaluation.completed_rows,

            "total_rows": evaluation.total_rows,

            "created_at": str(evaluation.created_at)

        }

    finally:

        db.close()


###########################################################
# Delete Evaluation
###########################################################

def remove_evaluation(

    user_id: int,

    evaluation_id: int

):

    db = SessionLocal()

    try:

        evaluation = get_evaluation(

            db,

            user_id,

            evaluation_id

        )

        if evaluation is None:

            return {

                "status": "error",

                "message": "Evaluation not found"

            }

        if (

            evaluation.results_path

            and

            Path(evaluation.results_path).exists()

        ):

            Path(

                evaluation.results_path

            ).unlink()

        delete_evaluation(

            db,

            evaluation

        )

        return {

            "status": "success"

        }

    finally:

        db.close()

###########################################################
# Get Evaluation Results
###########################################################

def get_evaluation_results(

    user_id: int,

    evaluation_id: int

):

    db = SessionLocal()

    try:

        ###################################################
        # Load Evaluation
        ###################################################

        evaluation = get_evaluation(

            db,

            user_id,

            evaluation_id

        )

        if evaluation is None:

            return {

                "status": "error",

                "message": "Evaluation not found"

            }

        ###################################################
        # Check File Exists
        ###################################################

        results_file = Path(

            evaluation.results_path

        )

        if not results_file.exists():

            return {

                "status": "error",

                "message": "Results file not found"

            }

        ###################################################
        # Load Workbook
        ###################################################

        workbook = load_workbook(

            results_file,

            data_only=True

        )

        worksheet = workbook.active

        rows = list(

            worksheet.iter_rows(

                values_only=True

            )

        )

        ###################################################
        # Empty Workbook
        ###################################################

        if not rows:

            workbook.close()

            return {

                "headers": [],

                "rows": []

            }

        ###################################################
        # Headers
        ###################################################

        headers = [

            "" if value is None else str(value)

            for value in rows[0]

        ]

        ###################################################
        # Data
        ###################################################

        data = []

        for row in rows[1:]:

            data.append(

                [

                    "" if value is None else value

                    for value in row

                ]

            )

        workbook.close()

        ###################################################
        # Response
        ###################################################

        return {

            "headers": headers,

            "rows": data

        }

    finally:

        db.close()

###########################################################
# Download Evaluation
###########################################################

def download_evaluation(

    user_id: int,

    evaluation_id: int

):

    db = SessionLocal()

    try:

        evaluation = get_evaluation(

            db,

            user_id,

            evaluation_id

        )

        if evaluation is None:

            return None

        results_file = Path(

            evaluation.results_path

        )

        if not results_file.exists():

            return None

        return str(

            results_file

        )

    finally:

        db.close()